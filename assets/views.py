from django.http import HttpResponse
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from complaints.models import Complaint

from .models import Asset, Department, Building
from .forms import AssetForm, DepartmentForm, BuildingForm


@never_cache
@login_required(login_url='login')
def asset_search(request):

    departments = Department.objects.all()
    buildings = Building.objects.all()

    assets = Asset.objects.all()

    department_id = request.GET.get('department')
    building_id = request.GET.get('building')
    item_id = request.GET.get('item_id')
    item_name = request.GET.get('item_name')

    if department_id:
        assets = assets.filter(
            department_id=department_id
        )

    if building_id:
        assets = assets.filter(
            building_id=building_id
        )

    if item_id:
        assets = assets.filter(
            item_id__icontains=item_id
        )

    if item_name:
        assets = assets.filter(
            item_name__icontains=item_name
        )

    return render(
        request,
        'assets/search.html',
        {
            'departments': departments,
            'buildings': buildings,
            'assets': assets
        }
    )


@never_cache
@login_required(login_url='login')
def asset_detail(request, asset_id):

    asset = get_object_or_404(
        Asset,
        id=asset_id
    )

    complaints = Complaint.objects.filter(
        asset=asset
    ).order_by('-created_at')

    return render(
        request,
        'assets/detail.html',
        {
            'asset': asset,
            'complaints': complaints
        }
    )


@never_cache
@login_required(login_url='login')
def asset_list(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    assets = Asset.objects.all().order_by('item_name')

    return render(
        request,
        'assets/admin_list.html',
        {
            'assets': assets
        }
    )


@never_cache
@login_required(login_url='login')
def add_asset(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    if request.method == 'POST':

        form = AssetForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect('asset_list')

    else:

        form = AssetForm()

    return render(
        request,
        'assets/add_asset.html',
        {
            'form': form
        }
    )
@never_cache
@login_required(login_url='login')
def edit_asset(request, asset_id):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    asset = get_object_or_404(
        Asset,
        id=asset_id
    )

    if request.method == 'POST':

        form = AssetForm(
            request.POST,
            instance=asset
        )

        if form.is_valid():

            form.save()

            return redirect('asset_list')

    else:

        form = AssetForm(
            instance=asset
        )

    return render(
        request,
        'assets/edit_asset.html',
        {
            'form': form,
            'asset': asset
        }
    )


@never_cache
@login_required(login_url='login')
def delete_asset(request, asset_id):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    asset = Asset.objects.get(id=asset_id)

    asset.delete()

    return redirect('asset_list')


@never_cache
@login_required(login_url='login')
def department_list(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    departments = Department.objects.all().order_by('name')

    return render(
        request,
        'assets/departments.html',
        {
            'departments': departments
        }
    )


@never_cache
@login_required(login_url='login')
def add_department(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    if request.method == 'POST':

        form = DepartmentForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect('department_list')

    else:

        form = DepartmentForm()

    return render(
        request,
        'assets/add_department.html',
        {
            'form': form
        }
    )


@never_cache
@login_required(login_url='login')
def delete_department(request, department_id):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    department = Department.objects.get(
        id=department_id
    )

    department.delete()

    return redirect('department_list')


@never_cache
@login_required(login_url='login')
def building_list(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    buildings = Building.objects.all().order_by('name')

    return render(
        request,
        'assets/buildings.html',
        {
            'buildings': buildings
        }
    )


@never_cache
@login_required(login_url='login')
def add_building(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    if request.method == 'POST':

        form = BuildingForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect('building_list')

    else:

        form = BuildingForm()

    return render(
        request,
        'assets/add_building.html',
        {
            'form': form
        }
    )


@never_cache
@login_required(login_url='login')
def delete_building(request, building_id):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    building = Building.objects.get(
        id=building_id
    )

    building.delete()

    return redirect('building_list')
@never_cache
@login_required(login_url='login')
def export_assets_excel(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Assets"

    headers = [
        'Asset ID',
        'Asset Name',
        'Category',
        'Department',
        'Building'
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    assets = Asset.objects.all()

    for row_num, asset in enumerate(
        assets,
        2
    ):

        sheet.cell(
            row=row_num,
            column=1
        ).value = asset.item_id

        sheet.cell(
            row=row_num,
            column=2
        ).value = asset.item_name

        sheet.cell(
            row=row_num,
            column=3
        ).value = asset.category

        sheet.cell(
            row=row_num,
            column=4
        ).value = str(asset.department)

        sheet.cell(
            row=row_num,
            column=5
        ).value = str(asset.building)

    response = HttpResponse(
        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename=assets.xlsx'

    workbook.save(response)

    return response