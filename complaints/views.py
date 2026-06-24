from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from .forms import ComplaintForm
from .models import Complaint
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.cache import never_cache
from .models import Complaint
from .forms import ComplaintForm, ComplaintStatusForm


@login_required
def complaint_create(request):

    if request.method == 'POST':
        form = ComplaintForm(request.POST, request.FILES)

        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.user = request.user
            complaint.save()

            return redirect('my_complaints')

    else:
        form = ComplaintForm()

    return render(
        request,
        'complaints/create.html',
        {'form': form}
    )


@login_required
def my_complaints(request):

    complaints = Complaint.objects.filter(
        user=request.user
    )

    return render(
        request,
        'complaints/list.html',
        {'complaints': complaints}
    )
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def admin_complaints(request):

    complaints = Complaint.objects.all().order_by(
        '-created_at'
    )

    return render(
        request,
        'complaints/admin_list.html',
        {
            'complaints': complaints
        }
    )

@staff_member_required
def complaint_list(request):

    complaints = Complaint.objects.all().order_by('-created_at')

    return render(
        request,
        'complaints/admin_list.html',
        {
            'complaints': complaints
        }
    )


@staff_member_required
def update_status(request, complaint_id):

    complaint = get_object_or_404(
        Complaint,
        id=complaint_id
    )

    if request.method == 'POST':

        form = ComplaintStatusForm(
            request.POST,
            instance=complaint
        )

        if form.is_valid():

            form.save()

            return redirect(
                'complaint_list'
            )

    else:

        form = ComplaintStatusForm(
            instance=complaint
        )

    return render(
        request,
        'complaints/update_status.html',
        {
            'form': form,
            'complaint': complaint
        }
    )
@never_cache
@login_required(login_url='/login/')
def export_complaints_excel(request):

    if not request.user.is_superuser:
        return redirect('user_dashboard')

    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Complaints"

    headers = [
        'Employee',
        'Asset',
        'Description',
        'Status',
        'Date'
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    complaints = Complaint.objects.all()

    for row_num, complaint in enumerate(
        complaints,
        2
    ):

        sheet.cell(
            row=row_num,
            column=1
        ).value = complaint.user.first_name

        sheet.cell(
            row=row_num,
            column=2
        ).value = complaint.asset.item_name

        sheet.cell(
            row=row_num,
            column=3
        ).value = complaint.description

        sheet.cell(
            row=row_num,
            column=4
        ).value = complaint.status

        sheet.cell(
            row=row_num,
            column=5
        ).value = str(complaint.created_at)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename=complaints.xlsx'

    workbook.save(response)

    return response